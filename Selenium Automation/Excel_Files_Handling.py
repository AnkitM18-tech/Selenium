from selenium import webdriver
from time import sleep
import openpyxl

"""path="C:/Users/OCAC/Desktop/Selenium/chromedriver.exe"
driver=webdriver.Chrome(path)"""

#Reading Excel File
excel_path="C:/Users/OCAC/Desktop/Selenium/BookStore.xlsx"
workbook=openpyxl.load_workbook(excel_path)         #getting the excel data
sheet=workbook.active

"""
#to get no of rows and no of columns
print(sheet.max_row)
print(sheet.max_column)"""

n_rows,n_cols=sheet.max_row,sheet.max_column
"""
#for loops to obtain values
for r in range(1,n_rows+1):
    for c in range(1,n_cols+1):
        print(sheet.cell(row=r,column=c).value,end="    ")
    print()
"""
#Writing Data into Excel file
#if you have multiple sheets then, workbook.get_sheet_by_name("sheet name")
for c in range(1,n_cols+1):
    sheet.cell(row=102,column=c).value="Selenium"        #add "Selenium" to every column of row=102.

workbook.save(excel_path)
from selenium import webdriver
from time import sleep
import openpyxl

path="C:/Users/OCAC/Desktop/Selenium/chromedriver.exe"
driver=webdriver.Chrome(path)

driver.get("https://courses.rifinder.com/dashboard/")
driver.maximize_window()
sleep(3)

excel_path="C:/Users/OCAC/Desktop/Selenium/Login.xlsx"
workbook=openpyxl.load_workbook(excel_path)         #getting the excel data
sheet=workbook.active

user_name=driver.find_element_by_xpath("//*[@id='user_login']")
password=driver.find_element_by_xpath('//*[@id="user_pass"]')
login=driver.find_element_by_xpath('//*[@id="wp-submit"]')

n_rows=sheet.max_row
for r in range(2,n_rows+1):
    user_data=sheet.cell(row=r,column=1).value
    pass_data=sheet.cell(row=r,column=2).value

    user_name.clear()
    password.clear()
    user_name.send_keys(user_data)
    password.send_keys(pass_data)
    login.click()
    #Validation
    if driver.title == "Dashboard - RiFinder":
        print("Test case Failed!")
        sheet.cell(row=r,column=3).value="Failed"
        workbook.save(excel_path)
    else:
        print("Test case Passed")
        sheet.cell(row=r,column=3).value="Passed"
        workbook.save(excel_path)
        break
workbook.close()
driver.close()